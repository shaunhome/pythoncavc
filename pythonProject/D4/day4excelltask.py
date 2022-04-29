import openpyxl

wb = openpyxl.load_workbook("gdp.xlsx")

wb.create_sheet("LA&C")
ns = wb["LA&C"]
ns["A1"] = "Latin American and Caribbean GDP"
ns["A3"] = "Country code"
ns["B3"] = "Country name"
ns["C3"] = "1990"
ns["D3"] = "2019"


metash = wb["Metadata - Countries"]
regs = metash.iter_rows(min_col=1, max_col=2, min_row=2, values_only=True)

dregs = dict(regs)
# print(dregs)
#for reg in regs:
#    print("{} : {}".format(reg[0], reg[1]))

sh = wb["Data"]
country_names = sh.iter_rows(min_col=1, max_col=64, min_row=5, values_only=True)
for cn in country_names:
    if cn[1] in dregs and dregs[cn[1]] == "Latin America & Caribbean":
        rc = dregs[cn[1]]
        print("{} : {} : {} : {}".format(cn[0], rc, cn[34], cn[63]))
        ns.append((cn[1], cn[0], cn[34], cn[63]))

wb.save("gdp_modified.xlsx")
wb.close()